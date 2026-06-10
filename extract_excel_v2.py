#!/usr/bin/env python3
"""提取企查查Excel — 只读能处理的，逐步保存"""
import json, os, re, time
import openpyxl

DATA_DIR = os.path.expanduser('~/文档/产业链/2025年最新产业链企业相关数据')
OUT_DIR = os.path.expanduser('~/industry_pdf_extracted')
os.makedirs(OUT_DIR, exist_ok=True)

SAVE_PATH = os.path.join(OUT_DIR, '_excel_summary.json')

# 记录已处理的
existing = {}
if os.path.exists(SAVE_PATH):
    with open(SAVE_PATH) as f:
        existing = json.load(f)

results = existing.copy()

for dir_name in sorted(os.listdir(DATA_DIR)):
    dir_path = os.path.join(DATA_DIR, dir_name)
    if not os.path.isdir(dir_path):
        continue
    
    chain_key = dir_name.replace('产业链企查查', '').replace('企业荣誉资质', '').replace('汇总', '').strip()
    if chain_key in results:
        print(f"[跳过] {chain_key} 已处理")
        continue
    
    xlsx_files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx') and not f.startswith('~')]
    if not xlsx_files:
        continue
    
    # 尝试每个xlsx
    for xlsx_file in xlsx_files:
        xlsx_path = os.path.join(dir_path, xlsx_file)
        print(f"[{chain_key}] {xlsx_file} ...")
        
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            
            # 读表头
            headers = None
            for row in ws.iter_rows(values_only=True):
                headers = [str(h).strip() if h else '' for h in row]
                break
            if not headers:
                wb.close()
                continue
            
            # 映射列
            col_map = {}
            for idx, h in enumerate(headers):
                if '公司名称' in h: col_map['company'] = idx
                elif '一级分类' in h: col_map['l1'] = idx
                elif '二级分类' in h: col_map['l2'] = idx
                elif '三级分类' in h: col_map['l3'] = idx
                elif '四级分类' in h: col_map['l4'] = idx
                elif '五级分类' in h: col_map['l5'] = idx
                elif '六级分类' in h: col_map['l6'] = idx
                elif '七级分类' in h: col_map['l7'] = idx
                elif '八级分类' in h: col_map['l8'] = idx
                elif '产业链节点' in h: col_map['node'] = idx
                elif '产业位置' in h: col_map['position'] = idx
                elif '融资上市' in h: col_map['listing'] = idx
            
            # 必须要有公司名+至少一个层级
            if 'company' not in col_map or 'l1' not in col_map:
                print(f"  跳过: 缺少必要列 {list(col_map.keys())}")
                wb.close()
                continue
            
            companies = []
            chain_paths = set()
            row_count = 0
            listed_count = 0
            
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                company_idx = col_map['company']
                if company_idx >= len(row) or not row[company_idx]:
                    continue
                company = str(row[company_idx]).strip()
                if not company or company == '公司名称':
                    continue
                
                # 构建层级路径
                levels = []
                for l in ['l1','l2','l3','l4','l5','l6','l7','l8']:
                    if l in col_map:
                        v = str(row[col_map[l]]).strip() if col_map[l] < len(row) and row[col_map[l]] else ''
                        if v and v != 'None' and v != '': 
                            levels.append(v)
                        else:
                            break
                    else:
                        break
                path = '>'.join(levels) if levels else chain_key
                if path:
                    chain_paths.add(path)
                
                position = str(row[col_map['position']]).strip() if 'position' in col_map and col_map['position'] < len(row) and row[col_map['position']] else ''
                listing = str(row[col_map['listing']]).strip() if 'listing' in col_map and col_map['listing'] < len(row) and row[col_map['listing']] else ''
                node = str(row[col_map['node']]).strip() if 'node' in col_map and col_map['node'] < len(row) and row[col_map['node']] else ''
                
                is_listed = bool(re.search(r'[板股]|上市|挂牌', listing))
                if is_listed:
                    listed_count += 1
                
                companies.append({
                    'name': company,
                    'chain_path': path,
                    'node': node,
                    'position': position,
                    'listing': listing,
                    'is_listed': is_listed,
                })
                row_count += 1
                
                # 上市公司数据已经够了，限制总公司数量
                if row_count >= 5000:
                    break
            
            wb.close()
            
            if companies:
                results[chain_key] = {
                    'chains': sorted(list(chain_paths)),
                    'companies': companies[:5000],  # 只存前5000
                    'total_companies': row_count,
                    'listed_companies': listed_count,
                }
                print(f"  {row_count}行, {listed_count}上市, {len(chain_paths)}链节点 [已保存]")
            else:
                print(f"  无数据")
            
            # 每个处理完都保存
            with open(SAVE_PATH, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=1, default=str)
            
            break  # 只处理第一个xlsx
        
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

print(f"\n✅ 完成! 共{len(results)}个产业链")
total_comp = sum(r.get('total_companies',0) for r in results.values())
total_listed = sum(r.get('listed_companies',0) for r in results.values())
print(f" 总公司: {total_comp}")
print(f" 上市公司: {total_listed}")
for k, v in sorted(results.items()):
    print(f"  {k}: {v['total_companies']}公司/{v['listed_companies']}上市/{len(v['chains'])}链")
