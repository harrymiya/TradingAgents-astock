#!/usr/bin/env python3
"""提取2025年企查查Excel数据"""
import json, os, sys, re, time

DATA_DIR = os.path.expanduser('~/文档/产业链/2025年最新产业链企业相关数据')
OUT_DIR = os.path.expanduser('~/industry_pdf_extracted')
os.makedirs(OUT_DIR, exist_ok=True)

import openpyxl

results = {}
all_stats = []

for dir_name in sorted(os.listdir(DATA_DIR)):
    dir_path = os.path.join(DATA_DIR, dir_name)
    if not os.path.isdir(dir_path):
        continue
    
    xlsx_files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx') and not f.startswith('~')]
    if not xlsx_files:
        continue
    
    xlsx_path = os.path.join(dir_path, xlsx_files[0])
    chain_name = dir_name.replace('产业链企查查', '').strip()
    
    print(f"[{chain_name}] 读取 {xlsx_files[0]}...")
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        
        # 解析表头
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            headers = [str(h) if h else '' for h in row]
            break
        
        # 找到关键列索引
        col_map = {}
        for idx, h in enumerate(headers):
            h_clean = h.strip() if h else ''
            if '公司名称' in h_clean: col_map['company'] = idx
            elif '一级分类' in h_clean: col_map['l1'] = idx
            elif '二级分类' in h_clean: col_map['l2'] = idx
            elif '三级分类' in h_clean: col_map['l3'] = idx
            elif '四级分类' in h_clean: col_map['l4'] = idx
            elif '五级分类' in h_clean: col_map['l5'] = idx
            elif '六级分类' in h_clean: col_map['l6'] = idx
            elif '七级分类' in h_clean: col_map['l7'] = idx
            elif '八级分类' in h_clean: col_map['l8'] = idx
            elif '产业链节点' in h_clean: col_map['node'] = idx
            elif '产业位置' in h_clean: col_map['position'] = idx
            elif '融资上市' in h_clean: col_map['listing'] = idx
        
        companies = []
        chain_paths = set()
        listed_count = 0
        
        for row in ws.iter_rows(values_only=True):
            if not row or not row[col_map.get('company', 0)]:
                continue
            company = str(row[col_map['company']]).strip() if col_map.get('company', 0) < len(row) else ''
            if not company or company == '公司名称':
                continue
            
            # 构建层级路径
            levels = []
            for l in ['l1','l2','l3','l4','l5','l6','l7','l8']:
                if l in col_map:
                    v = str(row[col_map[l]]).strip() if col_map[l] < len(row) else ''
                    if v and v != 'None':
                        levels.append(v)
                    else:
                        break
                else:
                    break
            path = '>'.join(levels) if levels else chain_name
            chain_paths.add(path)
            
            position = str(row[col_map['position']]).strip() if col_map.get('position', 0) < len(row) and row[col_map['position']] else ''
            listing = str(row[col_map['listing']]).strip() if col_map.get('listing', 0) < len(row) and row[col_map['listing']] else ''
            node = str(row[col_map['node']]).strip() if col_map.get('node', 0) < len(row) and row[col_map['node']] else ''
            
            is_listed = bool(re.search(r'[板股]|上市', listing))
            if is_listed:
                listed_count += 1
            
            companies.append({
                'name': company,
                'chain_path': path,
                'node': node,
                'position': position,
                'listing': listing,
                'is_listed': is_listed,
            })
        
        wb.close()
        
        results[chain_name] = {
            'chains': sorted(list(chain_paths)),
            'companies': companies,
            'total_companies': len(companies),
            'listed_companies': listed_count,
        }
        print(f"  {len(companies)}公司, {listed_count}上市, {len(chain_paths)}链节点")
        all_stats.append(f"{chain_name}: {len(companies)}公司/{listed_count}上市/{len(chain_paths)}链节点")
    
    except Exception as e:
        print(f"  ERROR: {e}")

# 保存
with open(os.path.join(OUT_DIR, '_excel_summary.json'), 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=1, default=str)

print("\n========== 汇总 ==========")
print(f"共{len(results)}个产业链")
for s in all_stats:
    print(f"  {s}")

# 计算总数
total_comp = sum(r['total_companies'] for r in results.values())
total_listed = sum(r['listed_companies'] for r in results.values())
print(f"\n总公司: {total_comp}")
print(f"上市公司: {total_listed}")
