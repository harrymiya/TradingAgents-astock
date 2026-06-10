#!/usr/bin/env python3
"""
批量提取2025年企查查Excel中的产业链数据
遍历 ~/文档/产业链/2025年最新产业链企业相关数据/ 下所有子目录，
提取包含14列的企查查产业链Excel，聚合输出到 ~/industry_pdf_extracted/_excel_summary.json
"""
import os
import json
import sys
from collections import OrderedDict

# ── paths ──
DATA_DIR = os.path.expanduser("~/文档/产业链/2025年最新产业链企业相关数据")
OUT_DIR = os.path.expanduser("~/industry_pdf_extracted")
OUT_FILE = os.path.join(OUT_DIR, "_excel_summary.json")

os.makedirs(OUT_DIR, exist_ok=True)

# ── helpers ──

def is_listing_info(val):
    """Check if a value looks like A-share listing info (contains 板/股/上市)."""
    if not val or not isinstance(val, str):
        return False
    val = val.strip()
    # A-share listing keywords
    keywords = ["板", "股", "上市"]
    return any(k in val for k in keywords)


def build_level_path(row_vals):
    """
    Build a path from level-1 to level-8 classification.
    row_vals[0..7] = 一级分类..八级分类
    """
    parts = []
    for i in range(8):
        v = row_vals[i]
        if v is not None and isinstance(v, str) and v.strip():
            parts.append(v.strip())
        else:
            break
    return ">".join(parts) if parts else ""


def classify_industry_key(level1_name):
    """
    Determine the top-level key for the output dict.
    Use the first available level-1 name, cleaned.
    """
    return level1_name.strip() if level1_name else ""


# ── main ──

def main():
    # Discover all .xlsx files
    xlsx_files = []
    for root, dirs, files in os.walk(DATA_DIR):
        for f in files:
            if f.endswith(".xlsx"):
                xlsx_files.append((root, f))

    # Sort for reproducibility
    xlsx_files.sort(key=lambda x: (x[0], x[1]))

    print(f"Found {len(xlsx_files)} .xlsx files in total.", flush=True)

    # Results dict: industry_key -> {"chains": set, "companies": list, "stats": {...}}
    results = OrderedDict()
    processed_count = 0
    skipped_not_14col = 0
    skipped_error = 0

    for idx, (root, filename) in enumerate(xlsx_files):
        rel_dir = os.path.relpath(root, DATA_DIR)
        filepath = os.path.join(root, filename)
        print(f"[{idx+1}/{len(xlsx_files)}] Processing: {rel_dir}/{filename}", flush=True)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

            # Find the first sheet that looks like a main data sheet
            sheet_names = wb.sheetnames
            # Prefer sheet with name containing "链" or starting with the directory name
            primary_sheet = None
            for sn in sheet_names:
                if "链" in sn:
                    primary_sheet = sn
                    break
            if primary_sheet is None and sheet_names:
                primary_sheet = sheet_names[0]
            elif primary_sheet is None:
                wb.close()
                skipped_error += 1
                continue

            ws = wb[primary_sheet]

            # Read header row to determine column layout
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

            # Check if this is a 14-column 企查查 file with 融资上市
            header_str = "|".join([str(h) if h else "" for h in header])

            if "融资上市" not in header_str:
                # Skip files without 融资上市 column
                wb.close()
                skipped_not_14col += 1
                continue

            # Map column indices (1-based)
            # Standard layout: cols 1-8 = 一级~八级分类, col 9 = 公司名称, col 11 = 产业链节点, col 12 = 产业位置, col 14 = 融资上市
            col_name = 9   # 公司名称
            col_node = 11  # 产业链节点
            col_pos = 12   # 产业位置
            col_list = 14  # 融资上市

            # Detect column header positions dynamically
            col_map = {}
            for i, h in enumerate(header):
                if h:
                    hs = str(h).strip()
                    if "公司名称" in hs or "企业名称" in hs:
                        col_map["name"] = i
                    elif "产业链节点" in hs:
                        col_map["node"] = i
                    elif "产业位置" in hs:
                        col_map["position"] = i
                    elif "融资上市" in hs:
                        col_map["listing"] = i

            # Fallback to standard positions if dynamic detection fails
            if "name" not in col_map:
                col_map["name"] = 8  # 0-based index 8 = col 9
            if "node" not in col_map:
                col_map["node"] = 10  # 0-based index 10 = col 11
            if "position" not in col_map:
                col_map["position"] = 11  # 0-based index 11 = col 12
            if "listing" not in col_map:
                col_map["listing"] = 13  # 0-based index 13 = col 14

            ci_name = col_map["name"]
            ci_node = col_map["node"]
            ci_pos = col_map["position"]
            ci_list = col_map["listing"]

            # Process data rows
            row_count = 0
            company_count = 0
            listed_count = 0
            chains_set = set()
            companies_list = []
            industry_key = None

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if row_count > 50000:  # Safety limit
                    break

                # Determine industry key from first non-None level-1 name
                level1 = row[0]
                if level1 and isinstance(level1, str) and level1.strip():
                    if industry_key is None:
                        industry_key = level1.strip()

                # Build level path
                path = build_level_path(row)
                if not path:
                    continue

                chains_set.add(path)

                # Company name
                cname = row[ci_name] if ci_name < len(row) else None
                if not cname or not isinstance(cname, str) or not cname.strip() or cname.strip() in ("-", ""):
                    continue

                cname = cname.strip()

                # Node
                node = row[ci_node] if ci_node < len(row) and row[ci_node] else ""
                node = str(node).strip() if isinstance(node, str) else str(node) if node else ""

                # Position
                position = row[ci_pos] if ci_pos < len(row) and row[ci_pos] else ""
                position = str(position).strip() if isinstance(position, str) else str(position) if position else ""

                # Listing info
                listing = row[ci_list] if ci_list < len(row) and row[ci_list] else ""
                listing = str(listing).strip() if isinstance(listing, str) else str(listing) if listing else ""

                # Determine if A-share listed
                is_listed = is_listing_info(listing)

                company_info = {
                    "name": cname,
                    "level": path,
                    "position": position,
                    "listing": listing,
                    "is_a_share_listed": is_listed
                }
                companies_list.append(company_info)
                company_count += 1
                if is_listed:
                    listed_count += 1

            wb.close()

            # Use directory name as key if industry_key not found
            if industry_key is None:
                industry_key = os.path.basename(root).replace("产业链企查查", "").replace("企查查", "").strip()
                if not industry_key:
                    industry_key = filename.replace(".xlsx", "")

            # Normalize key - remove trailing 产业链/企查查
            key = industry_key.replace("产业链", "").strip()

            # Store results
            if key not in results:
                results[key] = {
                    "chains": [],
                    "companies": [],
                    "stats": {
                        "total_companies": 0,
                        "a_share_listed_companies": 0,
                        "total_chains": 0
                    }
                }

            results[key]["chains"].extend(sorted(chains_set))
            results[key]["companies"].extend(companies_list)
            results[key]["stats"]["total_companies"] += company_count
            results[key]["stats"]["a_share_listed_companies"] += listed_count

            processed_count += 1
            print(f"  -> industry={key}, rows={row_count}, companies={company_count}, listed={listed_count}", flush=True)

        except Exception as e:
            skipped_error += 1
            print(f"  ERROR processing {rel_dir}/{filename}: {e}", flush=True)
            import traceback
            traceback.print_exc()

    # Deduplicate chains and companies across multiple files for same industry
    for key in results:
        results[key]["chains"] = sorted(set(results[key]["chains"]))
        # Deduplicate companies by name (keep first occurrence)
        seen_names = set()
        deduped = []
        for c in results[key]["companies"]:
            if c["name"] not in seen_names:
                seen_names.add(c["name"])
                deduped.append(c)
        results[key]["companies"] = deduped
        results[key]["stats"]["total_companies"] = len(deduped)
        results[key]["stats"]["total_chains"] = len(results[key]["chains"])

    # Write output
    output = OrderedDict()
    for key in sorted(results.keys()):
        r = results[key]
        output[key] = {
            "chains": r["chains"],
            "companies": [
                {
                    "name": c["name"],
                    "level": c["level"],
                    "position": c["position"],
                    "listing": c["listing"]
                }
                for c in r["companies"]
            ],
            "stats": r["stats"]
        }

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"Processing complete!")
    print(f"  Total xlsx files found: {len(xlsx_files)}")
    print(f"  14-col files processed: {processed_count}")
    print(f"  Skipped (no 融资上市): {skipped_not_14col}")
    print(f"  Errors: {skipped_error}")
    print(f"  Industries extracted: {len(results)}")
    print(f"  Output: {OUT_FILE}")
    print(f"{'='*60}")

    # Print industry stats
    total_companies = 0
    total_listed = 0
    print(f"\n{'Industry':<25} {'Companies':<12} {'A-Share Listed':<16} {'Chains':<8}")
    print("-" * 65)
    for key in sorted(results.keys()):
        r = results[key]
        s = r["stats"]
        print(f"{key:<25} {s['total_companies']:<12} {s['a_share_listed_companies']:<16} {s['total_chains']:<8}")
        total_companies += s["total_companies"]
        total_listed += s["a_share_listed_companies"]
    print("-" * 65)
    print(f"{'TOTAL':<25} {total_companies:<12} {total_listed:<16}")


if __name__ == "__main__":
    main()
